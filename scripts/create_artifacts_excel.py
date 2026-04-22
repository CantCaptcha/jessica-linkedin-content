#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Level 1 Artifacts'

headers = ['Domain', 'Control ID', 'Control Name', 'NIST Ref', 'Recommended Artifacts']

data = [
    ['AC', 'AC.1.001', 'Limit System Access', '3.1.1', 'Access Control Policy, User Access Matrix, AD User List, Onboarding Checklist, Account Creation Logs, Access Reviews, RBAC Matrix'],
    ['AC', 'AC.1.002', 'Limit Transactions & Functions', '3.1.2', 'Privilege Assignment Review, Admin Access Logs, Separation of Duties Matrix, User Permission Audit Reports'],
    ['AC', 'AC.1.003', 'Control Flow of CUI', '3.1.3', 'Data Flow Diagrams, Data Classification Policy, Secure Transmission Procedures (TLS 1.2+), Network Segmentation Diagrams'],
    ['AC', 'AC.1.004', 'Separate Duties', '3.1.4', 'SoD Matrix, Role Definitions, Conflict of Interest Policy, Approval Workflows, Audit Trail for Critical Transactions'],
    ['AT', 'AT.1.001', 'Ensure Personnel Trained', '3.2.1', 'Training Curriculum, Completion Records for All Personnel, Training Schedule, Role-Specific Materials, Acknowledgement of Policies'],
    ['AT', 'AT.1.002', 'Ensure Managers Trained', '3.2.2', 'Manager Training Records, Management-Level Security Awareness, Manager Responsibility Documentation, System Owner Training Certifications'],
    ['IR', 'IR.1.001', 'IR Policy & Procedures', '3.6.1', 'IR Policy (Signed/Approved), IR Procedures/Playbooks, Incident Classification Matrix, IRT Roster, Escalation Matrix, Reporting Procedures'],
    ['MA', 'MA.1.001', 'Maintenance Policy', '3.7.1', 'Maintenance Policy, Procedures Documentation, Schedule/Calendar, Personnel Access Documentation, Vendor Maintenance Agreements'],
    ['MA', 'MA.1.002', 'Maintenance Tools', '3.7.2', 'Approved Maintenance Tools Inventory, Tool Approval Process Documentation, Maintenance Tool Access Controls, Tool Usage Logs'],
    ['MA', 'MA.1.003', 'Maintenance Logs', '3.7.3', 'Maintenance Log Templates, Maintenance Activity Records, Log Review Procedures, Scheduled vs Unscheduled Documentation'],
    ['MA', 'MA.1.004', 'Maintenance Personnel', '3.7.4', 'Supervision Procedures, Personnel Background Check Records, Non-Disclosure Agreements, Escort/Chaperone Procedures, Access Logs with Signatures'],
    ['MP', 'MP.1.001', 'Access to Media', '3.8.1', 'Media Access Control Policy, Media Inventory (USB, External Drives), CUI Media Labeling Procedures, Media Checkout/Check-in Logs'],
    ['MP', 'MP.1.002', 'Media Sanitization', '3.8.2', 'Sanitization Policy (NIST SP 800-88), Sanitization Procedures, Method Inventory (Degauss, Wipe, Shred), Sanitization Logs Before Release'],
    ['MP', 'MP.1.003', 'Media Transport', '3.8.3', 'Media Transport Policy, Transport Chain-of-Custody Documentation, Encrypted Transport Procedures, Approved Carrier Agreements, Transport Logs'],
    ['MP', 'MP.1.004', 'Media Marking', '3.8.4', 'CUI Marking Policy/Procedures, Control Markings Guide (CUI, CUI//SP-SIMP), Marking Compliance Audit Reports, Document Templates with CUI'],
    ['PE', 'PE.1.001', 'Physical Access Policy', '3.9.1', 'Physical Security Policy, Access Control Zone Documentation, Badge Access Procedures, Visitor Access Procedures, Physical Access Logs (Entry/Exit)'],
    ['PE', 'PE.1.002', 'Access Authorization', '3.9.2', 'Access Authorization Matrix, Badge Issuance Procedures, Temporary Access Request Forms, Access Revocation Procedures, Personnel Clearance Documentation']
]

# Header style
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for col in range(1, 6):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = left_align
        cell.border = thin_border

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 80

wb.save('NIST-CMMC-Artifacts-by-Level.xlsx')
print("Excel file created with Level 1 artifacts!")
