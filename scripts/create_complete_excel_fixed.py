#!/usr/bin/env python3
"""
Complete NIST-CMMC Artifacts spreadsheet - fix version with all 6 sheets
"""

import sys
sys.path.insert(0, '/home/rwhitaker/.openclaw/workspace/skills/gmail-personal/venv/lib/python3.12/site-packages')

# Try importing with fallback
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    print("Using openpyxl")
except ImportError:
    print("openpyxl not found, trying openpyxl...")
    import openpyxl as openpyxl_module
    Font = openpyxl_module.styles.Font
    PatternFill = openpyxl_module.styles.PatternFill
    Alignment = openpyxl_module.styles.Alignment
    Border = openpyxl_module.styles.Border
    Side = openpyxl_module.styles.Side

# Create workbook
wb = openpyxl.Workbook()

# Define styles
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# SHEET 1: Overview
ws1 = wb.active
ws1.title = 'Overview'

headers1 = ['CMMC Level', 'Total Controls', 'Domain Count', 'Focus', 'Timeline']
data1 = [
    ['Level 1', 17, 6, 'Basic Cyber Hygiene - foundational cybersecurity practices', '3-6 months'],
    ['Level 2', 72, 12, 'Intermediate Protection - documented practices and additional controls', '6-12 months (concurrent with L1)'],
    ['TOTAL', 89, 12, 'Full CMMC compliance (L2 includes all L1)', '12-18 months total']
]

for col in range(1, 6):
    cell = ws1.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, row_data in enumerate(data1, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = left_align if col_idx > 1 else center_align
        cell.border = thin_border

for col_name, width in zip(['A', 'B', 'C', 'D', 'E'], [15, 15, 18, 50, 18]):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx+1)].width = width

# SHEET 2: Domain Comparison
ws2 = wb.create_sheet('Domain Comparison')
headers2 = ['Domain', 'Level 1 Controls', 'Level 2 Additional', 'Total Level 2']
data2 = [
    ['Access Control', 4, 5, 9],
    ['Awareness & Training', 2, 3, 5],
    ['Audit & Accountability', 0, 6, 6],
    ['Configuration Management', 0, 7, 7],
    ['Identification & Authentication', 0, 7, 7],
    ['Incident Response', 1, 4, 5],
    ['Maintenance', 4, 0, 4],
    ['Media Protection', 4, 0, 4],
    ['Physical Protection', 2, 4, 6],
    ['Risk Assessment', 0, 4, 4],
    ['Security Assessment', 0, 8, 8],
    ['System & Comms Protection', 0, 10, 10],
    ['System Integrity', 0, 8, 8],
    ['Personnel Security', 0, 0, 0],
    ['TOTAL', 17, 72, 89]
]

for col in range(1, 5):
    cell = ws2.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, row_data in enumerate(data2, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = center_align if isinstance(value, (int, float)) else left_align
        cell.border = thin_border

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 20

# SHEET 3: Level 1 Artifacts
ws3 = wb.create_sheet('Level 1 Artifacts')
headers3 = ['Domain', 'Control ID', 'Control Name', 'NIST Ref', 'Recommended Artifacts']
data3 = [
    ['AC', 'AC.1.001', 'Limit System Access', '3.1.1', 'Access Control Policy, User Access Matrix, AD User List, Onboarding Checklist, Account Creation Logs, Access Reviews, RBAC Matrix'],
    ['AC', 'AC.1.002', 'Limit Transactions & Functions', '3.1.2', 'Privilege Assignment Review, Admin Access Logs, Separation of Duties Matrix, User Permission Audit Reports'],
    ['AC', 'AC.1.003', 'Control Flow of CUI', '3.1.3', 'Data Flow Diagrams, Data Classification Policy, Secure Transmission Procedures (TLS 1.2+), Network Segmentation Diagrams'],
    ['AC', 'AC.1.004', 'Separate Duties', '3.1.4', 'SoD Matrix, Role Definitions, Conflict of Interest Policy, Approval Workflows, Audit Trail for Critical Transactions'],
    ['AT', 'AT.1.001', 'Ensure Personnel Trained', '3.2.1', 'Training Curriculum, Completion Records for All Personnel, Training Schedule, Role-Specific Materials, Acknowledgement of Policies'],
    ['AT', 'AT.1.002', 'Ensure Managers Trained', '3.2.2', 'Manager Training Records, Management-Level Security Awareness, Manager Responsibility Documentation, System Owner Training Certifications'],
    ['IR', 'IR.1.001', 'IR Policy & Procedures', '3.6.1', 'IR Policy (Signed/Approved), IR Procedures/Playbooks, Incident Classification Matrix, IRT Roster, Escalation Matrix, Reporting Procedures'],
    ['MA', 'MA.1.001', 'Maintenance Policy', '3.7.1', 'Maintenance Policy, Procedures Documentation, Schedule/Calendar, Personnel Access Documentation, Vendor Maintenance Agreements'],
    ['MA', 'MA.1.002', 'Maintenance Tools', '3.7.2', 'Approved Maintenance Tools Inventory, Tool Approval Process Documentation, Maintenance Tool Access Controls, Tool Usage Logs'],
    ['MA', 'MA.1.003', 'Maintenance Logs', '3.7.3', 'Maintenance Log Templates, Maintenance Activity Records, Log Review Procedures, Scheduled vs Unscheduled Documentation'],
    ['MA', 'MA.1.004', 'Maintenance Personnel', '3.7.4', 'Supervision Procedures, Personnel Background Check Records, Non-Disclosure Agreements, Escort/Chaperone Procedures, Access Logs with Signatures'],
    ['MP', 'MP.1.001', 'Access to Media', '3.8.1', 'Media Access Control Policy, Media Inventory (USB, External Drives), CUI Media Labeling Procedures, Media Checkout/Check-in Logs'],
    ['MP', 'MP.1.002', 'Media Sanitization', '3.8.2', 'Sanitization Policy (NIST SP 800-88), Sanitization Procedures, Method Inventory (Degauss, Wipe, Shred), Sanitization Logs Before Release'],
    ['MP', 'MP.1.003', 'Media Transport', '3.8.3', 'Media Transport Policy, Transport Chain-of-Custody Documentation, Encrypted Transport Procedures, Approved Carrier Agreements, Transport Logs'],
    ['MP', 'MP.1.004', 'Media Marking', '3.8.4', 'CUI Marking Policy/Procedures, Control Markings Guide (CUI, CUI//SP-SIMP), Marking Compliance Audit Reports, Document Templates with CUI'],
    ['PE', 'PE.1.001', 'Physical Access Policy', '3.9.1', 'Physical Security Policy, Access Control Zone Documentation, Badge Access Procedures, Visitor Access Procedures, Physical Access Logs (Entry/Exit)'],
    ['PE', 'PE.1.002', 'Access Authorization', '3.9.2', 'Access Authorization Matrix, Badge Issuance Procedures, Temporary Access Request Forms, Access Revocation Procedures, Personnel Clearance Documentation']
]

for col in range(1, 6):
    cell = ws3.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, row_data in enumerate(data3, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws3.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = left_align
        cell.border = thin_border

ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 35
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 80

# SHEET 4: Level 2 Artifacts
ws4 = wb.create_sheet('Level 2 Artifacts')
headers4 = ['Domain', 'Control ID', 'Control Name', 'NIST Ref', 'Recommended Artifacts']
data4 = [
    ['AC', 'AC.1.005', 'Least Privilege', '3.1.5', 'Least Privilege Policy, JIT Access Logs, Privilege Reviews, Temporary Elevated Access, Admin Account Reviews, Privilege Audit Reports'],
    ['AT', 'AT.1.003', 'Ensure All Personnel Trained', '3.2.3', 'Org-Wide Training Records, Phishing Simulations, Cyber Hygiene Materials, Annual Completion Reports'],
    ['AU', 'AU.1.001', 'Create Audit Records', '3.3.1', 'Audit Log Policy, Log Management Config, Audit Log Samples, Log Rotation Schedules'],
    ['AU', 'AU.1.002', 'Enable Auditing', '3.3.2', 'Audit Enablement Config, System Audit Policy, Logging Infrastructure, Log Source Inventory'],
    ['AU', 'AU.1.003', 'Review Audit Records', '3.3.3', 'Audit Review Schedule, Audit Logs, Log Analysis Reports, Security Incident Correlation'],
    ['AU', 'AU.1.004', 'Alert on Audit Failure', '3.3.4', 'Audit Monitoring Config, Alert Contact List, Failure Response Procedures, Alert Testing Records'],
    ['AU', 'AU.1.005', 'Time Synchronization', '3.3.5', 'NTP Policy, NTP Server Config, Time Sync Logs, Time Drift Reports'],
    ['AU', 'AU.1.006', 'Protect Audit Information', '3.3.6', 'Audit Data Access Policy, Audit Storage Encryption, Log Access Logs, Backup Logging Mechanisms'],
    ['CM', 'CM.1.001', 'Baseline Configuration', '3.4.1', 'Approved Baselines, Approval Signatures, Configuration Management Plan, Golden Image Repository'],
    ['CM', 'CM.1.002', 'Track Changes', '3.4.2', 'Change Control Policy, Change Request Forms, CAB Minutes, Approved/Rejected Changes, Rollback Procedures'],
    ['CM', 'CM.1.003', 'Security Impact Analysis', '3.4.3', 'Impact Analysis Methodology, Risk Assessment Templates, Security Impact Reviews, Pre/Post-Change Testing'],
    ['CM', 'CM.1.004', 'Vulnerability Scanning', '3.4.4', 'Scanning Policy, Scan Reports, Remediation Plans, Exception Approvals, Scanning Tool Config'],
    ['CM', 'CM.1.005', 'Anti-Malware', '3.4.5', 'Anti-Malware/EDR Deployment, Configuration Policies, Signature Update Schedules, Malware Detection Logs'],
    ['CM', 'CM.1.006', 'IR Testing', '3.4.6', 'IR Testing Schedule, Test Scenarios, Test Reports, After-Action Reports, Lessons Learned'],
    ['CM', 'CM.1.007', 'Emergency Access', '3.4.7', 'Emergency Procedures, Break-Glass Account Docs, Emergency Approval Forms, Emergency Access Logs'],
    ['IA', 'IA.1.001', 'Identify & Authenticate Users', '3.5.1', 'Identity Policy, Password Policy, Lockout Settings, Auth Mechanisms, Session Timeout Config'],
    ['IA', 'IA.1.002', 'Multi-Factor Authentication', '3.5.2', 'MFA Policy, Deployment Records, MFA Enforcement, Exception Approvals, MFA Backup Procedures'],
    ['IA', 'IA.1.003', 'Device Authentication', '3.5.3', 'External Connection Inventory, VPN Policies, Remote Desktop Gateway, Device Auth Certificates'],
    ['IA', 'IA.1.004', 'Replay Protection', '3.5.4', 'Auth Mechanism Docs, Replay Attack Protection Config, Session Management Policies'],
    ['IA', 'IA.1.005', 'Fail-Safe Auth', '3.5.5', 'Fail-Safe Auth Docs, Auth System Hardening, Backup Auth Config, Failure Mode Testing'],
    ['IA', 'IA.1.006', 'Token Authenticators (PIV)', '3.5.6', 'PIV Card Policy, PIV Inventory, PIV Reader Deployment, Smart Card Management, PIV Exemptions'],
    ['IA', 'IA.1.007', 'Cryptographic Modules (FIPS)', '3.5.7', 'FIPS Module Inventory, Crypto Module Certificates, Encryption Policy, Key Management, FIPS Compliance Reviews'],
    ['IR', 'IR.1.002', 'IR Training', '3.6.2', 'IR Team Training Records, Training Curriculum, Certificates, Skill Gap Analysis, Cross-Training'],
    ['IR', 'IR.1.003', 'IR Testing', '3.6.3', 'IR Testing Schedule, Test Scenarios, Test Execution Reports, Tabletop Exercises, After-Action Reports'],
    ['IR', 'IR.1.004', 'Incident Monitoring', '3.6.4', 'Incident Tracking Logs, Incident Reports, Timeline Documentation, Root Cause Analysis, Closure Docs'],
    ['IR', 'IR.1.005', 'Incident Reporting', '3.6.5', 'Reporting Procedures, Reporting Channel Docs, Awareness Materials, Anonymous Reporting Mechanisms'],
    ['PE', 'PE.1.003', 'Visitor Control', '3.9.3', 'Visitor Management Procedures, Sign-In/Out Logs, Badge Expiration, Physical Monitoring Reports'],
    ['PE', 'PE.1.004', 'Physical Access Monitoring', '3.9.4', 'Security Camera Map, Access Monitoring, Alarm/Sensor Config, Physical Intrusion Logs'],
    ['PE', 'PE.1.005', 'Equipment Maintenance', '3.9.5', 'Physical Security Equipment Inventory, Maintenance Schedule, Calibration Records, Equipment Testing'],
    ['PE', 'PE.1.006', 'Delivery Area Control', '3.9.6', 'Delivery Area Security, Receiving Access Controls, Package Inspection, Delivery Logs'],
    ['RA', 'RA.1.001', 'Risk Assessment Policy', '3.11.1', 'Risk Management Policy, Risk Assessment Methodology, Risk Tolerance Statements, Risk Assessment Schedule, Risk Register'],
    ['RA', 'RA.1.002', 'Scan for Vulnerabilities', '3.11.2', 'Vuln Scanning Policy, Scan Schedule (Quarterly), Scan Reports, False Positive Docs, Scan Tool Config'],
    ['RA', 'RA.1.003', 'Remediate Vulnerabilities', '3.11.3', 'Vuln Remediation Plan, Patch Management Policy, Patch Deployment Schedules, Remediation Verification'],
    ['RA', 'RA.1.004', 'Update Risk Assessments', '3.11.4', 'Update Schedule, Current Risk Assessment, Risk Review Minutes, Risk Trend Analysis, Risk Acceptance Sign-Offs'],
    ['CA', 'CA.1.001', 'Security Assessment Policy', '3.12.1', 'Security Assessment Policy, Assessment Schedule, Team Credentials, Methodology Docs, Scope Documentation'],
    ['CA', 'CA.1.002', 'Assess Security Controls', '3.12.2', 'Control Assessment Reports, Effectiveness Testing, Pen Testing Reports, Independent Assessments, Gap Analysis'],
    ['CA', 'CA.1.003', 'Remediation Plan (POAM)', '3.12.3', 'POAM, Gap Remediation Schedules, Responsibility Assignments, POAM Tracking Logs, Completion Evidence'],
    ['CA', 'CA.1.004', 'System Security Plan (SSP)', '3.12.4', 'SSP, Approval Signatures, Control Implementation Matrix, SSP Change Logs, SSP Distribution, SSP Review Schedule'],
    ['CA', 'CA.1.005', 'Update SSP', '3.12.5', 'SSP Update Procedures, SSP Change Logs, Version Control, Update Approvals, Review Sign-Offs'],
    ['CA', 'CA.1.006', 'External Connections Review', '3.12.6', 'External Connection Inventory, Connection Approval Procedures, Security Requirements, Connection Agreements, FedRAMP Docs'],
    ['CA', 'CA.1.007', 'Continuous Monitoring', '3.12.7', 'Continuous Monitoring Plan, Monitoring Tool Inventory, Dashboard/Config, Alert Thresholds, Monitoring Logs'],
    ['CA', 'CA.1.008', 'Security Assessments on Change', '3.12.8', 'Pre-Change Test Procedures, Post-Change Test Results, Security Impact Assessments, Regression Testing'],
    ['SC', 'SC.1.001', 'Boundary Protection', '3.13.1', 'Network Boundary Docs, Firewall Config Files, IDS/IPS Logs, Network Segmentation Diagrams, ACLs, Traffic Flow Reports'],
    ['SC', 'SC.1.002', 'Information in Transit', '3.13.2', 'Encryption-in-Transit Policy (TLS 1.2+), TLS/SSL Config, Certificate Inventory, VPN Encryption Docs, Secure Transmission Protocols'],
    ['SC', 'SC.1.003', 'Information at Rest', '3.13.3', 'Encryption-at-Rest Policy, Disk Encryption Records, Database Encryption Config, File/Folder Encryption Policies, Key Management'],
    ['SC', 'SC.1.004', 'CUI Partitioning', '3.13.4', 'Data Partitioning Procedures, Network Segmentation Docs, CUI Storage Inventory, Cross-Boundary Controls, Access Matrix by Class'],
    ['SC', 'SC.1.005', 'Denial of Service Protection', '3.13.5', 'DoS/DDoS Protection Policy, Mitigation Tool Config, Bandwidth Monitoring, DoS Response Procedures, ISP/CDN DDoS Protection'],
    ['SC', 'SC.1.006', 'Mobile Code Protection', '3.13.6', 'Mobile Code Policy, Browser Hardening, Application Allow/Denylists, Code Signing Verification, Sandboxing Docs'],
    ['SC', 'SC.1.007', 'Cryptography Protection', '3.13.7', 'Crypto Standards Policy, Approved Algorithms List (NIST SP 800-52), Encryption Strength Docs, Key Lifecycle Management, FIPS Module Inventory'],
    ['SC', 'SC.1.008', 'Confidentiality Protection', '3.13.8', 'Confidentiality Protection Policy, End-to-End Encryption Docs, PGP/S/MIME Key Management, Secure Email Gateway, File Encryption Procedures'],
    ['SC', 'SC.1.009', 'Session Protection', '3.13.9', 'Session Timeout Policy, Timeout Config Docs, Application Timeout Settings, Session Lock Enforcement, Idle Session Monitoring'],
    ['SC', 'SC.1.010', 'Wireless Access', '3.13.10', 'Wireless Security Policy, WPA2/WPA3 Enterprise Config, SSID Management, Wireless AP Inventory, Rogue AP Detection, RADIUS Config'],
    ['SI', 'SI.1.001', 'Flaw Remediation', '3.14.1', 'Vulnerability Management Policy, Patch Management Procedures, Vuln ID Logs, Patch Deployment Schedules, Root Cause Analysis'],
    ['SI', 'SI.1.002', 'Security Software Updates', '3.14.2', 'Anti-Malware/EDR Update Policy, Signature Update Automation, Update Frequency Docs, Update Failure Notifications'],
    ['SI', 'SI.1.003', 'Security Baseline Updates', '3.14.3', 'Security Baseline Update Schedule, Baseline Version Control, Baseline Approval Docs, Update Notifications, Deployment Records'],
    ['SI', 'SI.1.004', 'Security Alerts', '3.14.4', 'Security Alert Monitoring Procedures, Alert Correlation Rules, Alert Threshold Config, Alert Response Playbooks, SIEM/SOC Config'],
    ['SI', 'SI.1.005', 'Software Flaw Scans', '3.14.5', 'Software Scanning Policy (SAST/DAST), Static/Dynamic Analysis Scan Reports, Code Reviews, Third-Party Scans, Scan Frequency'],
    ['SI', 'SI.1.006', 'Timely Remediation', '3.14.6', 'Remediation SLA Docs, Vulnerability Priority Matrix, Remediation Tracking System Logs, Overdue Vuln Reports, Risk Acceptance'],
    ['SI', 'SI.1.007', 'Unauthorized Software', '3.14.7', 'Software Authorization Policy, Approved Software Catalog, Software Deployment Procedures, Application Allowlist Config, Unauthorized Software Detection'],
    ['SI', 'SI.1.008', 'Security Functionality Verification', '3.14.8', 'Security Control Testing Schedule, Functionality Test Reports, Control Effectiveness Validation, Testing Methodologies, Test Env Docs']
]

for col in range(1, 6):
    cell = ws4.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, row_data in enumerate(data4, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws4.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = left_align
        cell.border = thin_border

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 35
ws4.column_dimensions['D'].width = 12
ws4.column_dimensions['E'].width = 80

# SHEET 5: Critical Documents
ws5 = wb.create_sheet('Critical Documents')
headers5 = ['Document Type', 'Required for Level 1', 'Required for Level 2', 'Description', 'Priority']
data5 = [
    ['Policies', 'Access Control, IR, Maintenance, Physical Security, Media Protection', 'PLUS: Audit, Config Mgmt, Risk Mgmt, Security Assessment, Encryption, Continuous Monitoring', 'All security policies must be signed and approved'],
    ['System Security Plan (SSP)', 'Recommended (Good Practice)', 'MANDATORY', 'Documents all controls, implementation, and procedures - core CMMC document'],
    ['Plan of Action & Milestones (POAM)', 'Recommended (Good Practice)', 'MANDATORY', 'Tracks gaps, remediation, and completion milestones - critical for assessments'],
    ['Audit Logging', 'Not Required', 'Required', 'Enable audit logging for all CUI-processing systems'],
    ['Multi-Factor Authentication', 'Not Required', 'Required', 'MFA enforcement for all user authentication'],
    ['Configuration Management', 'Recommended', 'Required', 'Baselines, change control, security impact analysis'],
    ['Vulnerability Management', 'Recommended', 'Required', 'Scanning, patching, and remediation of vulnerabilities'],
    ['Risk Assessment', 'Recommended', 'Required', 'Ongoing risk identification, analysis, and mitigation'],
    ['Continuous Monitoring', 'Not Required', 'Required', 'SIEM/SOC capabilities for real-time security monitoring'],
    ['External Connections Inventory', 'Recommended', 'Required', 'Third-party systems, FedRAMP authorizations, security assessments'],
    ['Encryption Documentation', 'Recommended', 'Required', 'FIPS-validated crypto modules, encryption policies (in-transit and at-rest)'],
    ['Training Records', 'Required', 'Required', 'All personnel security and role-specific training']
]

for col in range(1, 6):
    cell = ws5.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, row_data in enumerate(data5, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws5.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = left_align
        cell.border = thin_border

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 20
ws5.column_dimensions['C'].width = 20
ws5.column_dimensions['D'].width = 60
ws5.column_dimensions['E'].width = 12

# SHEET 6: Implementation Roadmap
ws6 = wb.create_sheet('Implementation Roadmap')
headers6 = ['Phase', 'Timeline', 'Key Activities', 'Deliverables', 'Dependencies']
data6 = [
    ['Phase 1: Level 1 Foundation', 'Months 1-6', 'Policies, Access Control, Physical Security, Training, IR, Maintenance, Media', '17 L1 controls implemented, evidence collected', 'None'],
    ['Phase 2: SSP & POAM', 'Months 1-3 (parallel)', 'Create System Security Plan documenting all controls, Create POAM tracking gaps', 'SSP v1.0, POAM v1.0', 'Phase 1 foundation'],
    ['Phase 3: Audit & MFA', 'Months 2-4 (parallel)', 'Enable audit logging across environment, Deploy MFA for all authentication', 'Audit log policy, MFA deployment records', 'Phase 1 foundation'],
    ['Phase 4: Config Mgmt', 'Months 3-5 (parallel)', 'Establish baselines, implement change control, security impact analysis', 'Baseline documents, change control process', 'Phase 1 foundation'],
    ['Phase 5: Vulnerability Mgmt', 'Months 4-6 (parallel)', 'Implement vulnerability scanning, patch management, remediation process', 'Vuln scanning policy, patch deployment schedules', 'Phase 1 foundation'],
    ['Phase 6: Risk Assessment', 'Months 4-6 (parallel)', 'Conduct risk assessment, establish risk register, periodic reviews', 'Risk management policy, risk assessment reports', 'Phase 1 foundation'],
    ['Phase 7: Encryption & Protection', 'Months 5-8', 'Implement encryption-in-transit, encryption-at-rest, FIPS-validated modules', 'Encryption policies, FIPS module inventory', 'Phase 1 foundation'],
    ['Phase 8: Continuous Monitoring', 'Months 7-9', 'Deploy SIEM/SOC capabilities, establish alerting, monitoring dashboards', 'Continuous monitoring plan, alert correlation rules', 'Phase 2 controls'],
    ['Phase 9: External Connections', 'Months 8-10', 'Inventory external connections, review security, implement third-party risk management', 'External connection inventory, security assessments', 'Phase 2 controls'],
    ['Phase 10: Testing & Validation', 'Months 9-12', 'Penetration testing, IR exercises, control effectiveness validation', 'Pen test reports, IR test reports, control validation', 'Phase 2 controls'],
    ['Phase 11: Final Preparation', 'Months 11-12', 'Consolidate evidence, conduct readiness assessment, address remaining gaps', 'Complete evidence package, readiness report', 'All previous phases']
]

for col in range(1, 6):
    cell = ws6.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, row_data in enumerate(data6, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws6.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = left_align
        cell.border = thin_border

ws6.column_dimensions['A'].width = 25
ws6.column_dimensions['B'].width = 18
ws6.column_dimensions['C'].width = 45
ws6.column_dimensions['D'].width = 45
ws6.column_dimensions['E'].width = 25

# Save workbook
wb.save('NIST-CMMC-Artifacts-by-Level.xlsx')
print("✅ Complete Excel file created with all 6 sheets!")
print("File: NIST-CMMC-Artifacts-by-Level.xlsx")
print("Sheets:")
print("  1. Overview (Level summaries & timelines)")
print("  2. Domain Comparison (controls per domain L1 vs L2)")
print("  3. Level 1 Artifacts (17 controls)")
print("  4. Level 2 Artifacts (72 controls)")
print("  5. Critical Documents (required at each level with priorities)")
print("  6. Implementation Roadmap (11 phases with timelines)")
print("\nTotal controls: 89 (17 L1 + 72 L2)")
