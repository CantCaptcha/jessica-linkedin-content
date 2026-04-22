#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = openpyxl.Workbook()

# Define styles
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# SHEET 1: Overview
ws1 = wb.active
ws1.title = 'Overview'

headers1 = ['CMMC Level', 'Total Controls', 'Domain Count', 'Focus', 'Timeline']
data1 = [
    ['Level 1', 17, 6, 'Basic Cyber Hygiene - foundational cybersecurity practices', '3-6 months'],
    ['Level 2', 72, 12, 'Intermediate Protection - documented practices and additional controls', '6-12 months (concurrent with L1)'],
    ['TOTAL', 89, 12, 'Full CMMC compliance (L2 includes all L1)', '12-18 months total']
]

for col in range(1, 6):
    cell = ws1.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, row_data in enumerate(data1, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = left_align if col_idx > 1 else center_align
        cell.border = thin_border

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 50
ws1.column_dimensions['E'].width = 18

# SHEET 2: Domain Comparison
ws2 = wb.create_sheet('Domain Comparison')
headers2 = ['Domain', 'Level 1 Controls', 'Level 2 Additional', 'Total Level 2']
data2 = [
    ['Access Control', 4, 5, 9],
    ['Awareness & Training', 2, 3, 5],
    ['Audit & Accountability', 0, 6, 6],
    ['Configuration Management', 0, 7, 7],
    ['Identification & Authentication', 0, 7, 7],
    ['Incident Response', 1, 4, 5],
    ['Maintenance', 4, 0, 4],
    ['Media Protection', 4, 0, 4],
    ['Physical Protection', 2, 4, 6],
    ['Risk Assessment', 0, 4, 4],
    ['Security Assessment', 0, 8, 8],
    ['System & Comms Protection', 0, 10, 10],
    ['System Integrity', 0, 8, 8],
    ['Personnel Security', 0, 0, 0],
    ['TOTAL', 17, 72, 89]
]

for col in range(1, 6):
    cell = ws2.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, row_data in enumerate(data2, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = center_align if isinstance(value, (int, float)) else left_align
        cell.border = thin_border

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 20

# Save workbook
wb.save('NIST-CMMC-Artifacts-by-Level.xlsx')
print("Excel file created successfully!")
